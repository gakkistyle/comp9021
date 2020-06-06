# COMP9021 Practice 9 - Solutions


from collections import Counter, defaultdict

import openpyxl


class ElectionError(Exception):
    def __init__(self, message):
        self.message = message


class Election:
    def __init__(self, ballot):
        self.ballot = ballot
        wb = openpyxl.load_workbook(self.ballot)
        if len(wb.sheetnames) != 1:
            raise VotingError('Spreadsheet should have only one sheet')
        sheet = wb[wb.sheetnames[0]]
        if sheet.max_row < 3:
            raise VotingError('Spreadsheet should have at least three rows')
        if sheet.max_column < 2:
            raise VotingError('Spreadsheet should have at least two columns')
        if (sheet[1][0].value != 'Election results' and
                                    sheet[1][1: sheet.max_column] != [None] * (sheet.max_column + 1)
           ):
            raise VotingError('Improper spreadsheet title')
        if sheet[2][0].value != 'Number of votes':
            raise VotingError('Leftmost cell below spreadsheet title should be "Number of votes"')
        for column in range(1, sheet.max_column):
            if (not isinstance(sheet[2][column].value, str) or not sheet[2][column].value.istitle()
               ):
                raise VotingError('Except for the leftmost one, cells below spreadsheet title '
                                                                   'should all be capitalised names'
                                 )
        for row in range(3, sheet.max_row + 1):
            if (not isinstance(sheet[row][0].value, int) or sheet[row][0].value < 1):
                raise VotingError('Cells below "Number of votes" should all be '
                                                                        'strictly positive integers'
                                 )
        for row in range(3, sheet.max_row + 1):
            if (any(not isinstance(sheet[row][column].value, int)
                                                            for column in range(1, sheet.max_column)
                   ) or set(sheet[row][column].value for column in range(1, sheet.max_column)) !=
                                                                     set(range(1, sheet.max_column))
               ):
                raise VotingError('All vote results should be candidate rankings, '
                                                          'from 1 to the total number of candidates'
                                 )
        self.candidates = [sheet[2][column].value for column in range(1, sheet.max_column)]
        self.tallies = [sheet[row][0].value for row in range(3, sheet.max_row + 1)]
        self.candidate_rankings = [dict((self.candidates[column - 1], sheet[row][column].value)
                                                            for column in range(1, sheet.max_column)
                                       ) for row in range(3, sheet.max_row + 1)
                                  ]
        self.ranked_candidates = [dict((sheet[row][column].value, self.candidates[column - 1])
                                                            for column in range(1, sheet.max_column)
                                      ) for row in range(3, sheet.max_row + 1)
                                 ]
        self._field_width = max(len(candidate) for candidate in self.candidates)

    def __repr__(self):
        return (f"Voting('{self.ballot}')")

    def __str__(self):
        table ='Number of votes'
        for candidate in self.candidates:
            table += f'  {candidate:^{self._field_width}}'
        for i in range(len(self.tallies)):
            table += f'\n{self.tallies[i]:^15d}'
            table += ''.join(f'  {self.candidate_rankings[i][candidate]:^{self._field_width}d}'
                                                                    for candidate in self.candidates
                            )
        return table

    def _print_winners(self, winners):
        winners = sorted(winners)
        if not(len(winners)):
            print('There is no winner.')
        elif len(winners) == len(self.candidates):
            print('All candidates are winners.')
        elif len(winners) == 1:
            print(f'The winner is {winners[0]}.')
        else:
            print(f'The winners are {winners[0]}', end = '')
            for i in range(1, len(winners) - 1):
                print(f', {winners[i]}', end = '')
            print(f' and {winners[len(winners) - 1]}.')           

    def one_round_winners(self):
        self._print_winners(self._get_top_candidates())
                    
    def two_round_winners(self):
        first_round_results = self._get_top_candidates(at_least_two = True)
        self._print_winners(self._get_top_candidates_amongst(first_round_results))
        
    def elimination_winners(self):
        remaining_candidates = set(self.candidates)
        eliminated_candidates = set()
        while len(remaining_candidates) - len(eliminated_candidates):
            remaining_candidates -= eliminated_candidates
            eliminated_candidates = self._get_top_candidates_amongst(remaining_candidates,
                                                                                 to_eliminate = True
                                                                    )
        self._print_winners(remaining_candidates)

    def de_borda_winners(self):
        tallies_for_all_candidates = Counter()
        for candidate in self.candidates:
            tallies_for_all_candidates[candidate] = sum(
                                       self.tallies[i] * (6 - self.candidate_rankings[i][candidate])
                                                                   for i in range(len(self.tallies))
                                                         )
        self._print_winners(self._select_candidates(tallies_for_all_candidates.most_common()))

    def de_condorcet_winners(self):
        pairwise_contests = defaultdict(int)
        candidates = sorted(self.candidates)
        for a in range(len(candidates)):
            for b in range(a + 1, len(candidates)):
                pairwise_contests[candidates[a], candidates[b]] = sum(
                                               (2 * (self.candidate_rankings[i][candidates[b]] >
                                                           self.candidate_rankings[i][candidates[a]]
                                                    ) - 1
                                               ) * self.tallies[i] for i in range(len(self.tallies))
                                                                     )
        winners = set()
        for candidate in self.candidates:
            for contestant in self.candidates:
                if candidate < contestant:
                    if pairwise_contests[candidate, contestant] < 0:
                        break
                elif candidate > contestant:
                    if pairwise_contests[contestant, candidate] > 0:
                        break
            else:
                winners.add(candidate)
        self._print_winners(sorted(winners))

    def _get_candidates(self, selection_method, at_least_two = False, to_eliminate = False):
        return self._select_candidates(self._candidate_votes(selection_method),
                                            at_least_two = at_least_two, to_eliminate = to_eliminate
                                      )

    def _get_top_candidates(self, at_least_two = False):
        return self._get_candidates(lambda i: self.ranked_candidates[i][1], at_least_two)
                                                      
    def _get_top_candidates_amongst(self, candidates, to_eliminate = False):
        return self._get_candidates(
                       lambda i: self.ranked_candidates[i][min(self.candidate_rankings[i][candidate]
                                                                         for candidate in candidates
                                                              )
                                                          ], to_eliminate = to_eliminate
                                   )

    def _candidate_votes(self, selection_method):
        tallies_for_selected_candidates = Counter()
        for i in range(len(self.tallies)):
            tallies_for_selected_candidates[selection_method(i)] += self.tallies[i]
        return tallies_for_selected_candidates.most_common()
        
    def _select_candidates(self, results, *, at_least_two = False, to_eliminate = False):
        if to_eliminate:
            results = list(reversed(results))
        selected = {results[0][0]}
        nb_of_votes_for_selected = results[0][1]
        for (candidate, nb_of_votes) in results[1: ]:
            if nb_of_votes == nb_of_votes_for_selected:
                selected.add(candidate)
            elif len(selected) == 1 and at_least_two:
                selected.add(candidate)
                nb_of_votes_for_selected = nb_of_votes
            else:
                break
        return selected
                

if __name__ == '__main__':
    election = Election('election_1.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
    print('\n')
    
    election = Election('election_2.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
    print('\n')

    election = Election('election_3.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
    print('\n')

    election = Election('election_4.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
    print('\n')

    election = Election('election_5.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
    print('\n')

    election = Election('election_6.xlsx')
    print('Election results:')
    print(election)
    print()
    print('One round method:')
    election.one_round_winners()
    print()
    print('Two round method:')
    election.two_round_winners()
    print()
    print('Elimination method:')
    election.elimination_winners()
    print()
    print('De Borda method:')
    election.de_borda_winners()
    print()
    print('De Condorcet method:')
    election.de_condorcet_winners()
